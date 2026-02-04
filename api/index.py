from datetime import datetime
from io import BytesIO

from flask import Flask, render_template_string, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList

from auto_table_core import TEMPLATE, get_table_data

app = Flask(__name__)


def _build_workbook(rows, stats, selected_columns, include_stats, filters):
    """Build an XLSX workbook matching current table + optional stats/charts."""
    wb = Workbook()
    ws_table = wb.active
    ws_table.title = "Table"

    # Default columns if none selected
    default_columns = [
        "Region",
        "Province",
        "BEIS School ID",
        "Schedule",
        "Calendar Status",
        "Start Time",
        "End Time",
        "Installation Status",
        "Starlink Status",
        "Approval",
        "Blocker",
    ]
    columns = selected_columns or default_columns
    max_col = len(columns)

    # Header style
    header_font = Font(bold=True, color="020617")
    header_fill = PatternFill("solid", fgColor="CBD5F5")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Title row above the table
    title_cell = ws_table.cell(row=1, column=1, value="LEOxSOLAR Schedule Monitoring Report")
    ws_table.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell.font = Font(bold=True, size=14, color="0F172A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write header row (row 2)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws_table.cell(row=2, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    # Data rows
    # Row highlighting to match dashboard warning/critical colors
    warning_fill = PatternFill("solid", fgColor="FDE68A")  # yellow
    critical_fill = PatternFill("solid", fgColor="F97373")  # red
    band_even_fill = PatternFill("solid", fgColor="F9FAFB")  # banding for even rows
    today = datetime.now().date()

    for row_idx, row in enumerate(rows, start=3):
        star = (row.get("Starlink Status") or "").lower()
        appr = (row.get("Approval") or "").lower()
        row_fill = None
        if "declin" in appr:
            row_fill = critical_fill
        elif star != "activated":
            row_fill = warning_fill

        for col_idx, col_name in enumerate(columns, start=1):
            value = row.get(col_name, "")
            cell = ws_table.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            # Align Region / School / Blocker to left, others center
            if col_name in ("Region", "Province", "BEIS School ID", "Blocker", "Installation Status"):
                cell.alignment = left
            else:
                cell.alignment = center
            if row_fill is not None:
                cell.fill = row_fill
            elif row_idx % 2 == 0:
                cell.fill = band_even_fill

            # Schedule cell coloring: closer to today => darker red
            if col_name == "Schedule":
                sched_str = str(value).strip()
                if sched_str:
                    sched_fill = None
                    try:
                        sched_date = datetime.strptime(sched_str, "%b. %d, %Y").date()
                        delta_days = abs((sched_date - today).days)
                        if delta_days <= 1:
                            # very close: brightest red (matches dashboard red)
                            sched_fill = PatternFill("solid", fgColor="DC2626")
                        elif delta_days <= 3:
                            # near: medium bright red
                            sched_fill = PatternFill("solid", fgColor="F97373")
                        elif delta_days <= 7:
                            # upcoming: light red
                            sched_fill = PatternFill("solid", fgColor="FECACA")
                    except Exception:
                        sched_fill = None
                    if sched_fill is not None:
                        cell.fill = sched_fill

    # Auto-fit-ish column widths based on header length, with overrides per column
    width_overrides = {
        "Region": 14,
        "Province": 14,
        "BEIS School ID": 14,
        "Schedule": 16,
        "Calendar Status": 14,
        "Start Time": 10,
        "End Time": 10,
        "Installation Status": 26,
        "Starlink Status": 14,
        "Approval": 18,
        "Blocker": 40,
    }
    for col_idx, col_name in enumerate(columns, start=1):
        letter = ws_table.cell(row=2, column=col_idx).column_letter
        base_width = max(10, min(30, len(str(col_name)) + 4))
        width = width_overrides.get(col_name, base_width)
        ws_table.column_dimensions[letter].width = width

    # Freeze header row and first column
    ws_table.freeze_panes = "B3"

    if include_stats:
        ws_stats = wb.create_sheet(title="Summary")
        # Title for summary sheet
        ws_stats["A1"] = "LEOxSOLAR Summary"
        ws_stats["A1"].font = Font(bold=True, size=14, color="0F172A")
        ws_stats.merge_cells("A1:B1")

        # Filters block
        ws_stats["A2"] = "Filters"
        ws_stats["A2"].font = Font(bold=True, color="0F172A")
        filters_map = [
            ("Lot #", filters.get("lot") or "All"),
            ("Region", filters.get("region") or "All"),
            ("Schedule", filters.get("schedule") or "All"),
            ("Installation", filters.get("installation") or "All"),
            ("Tile", filters.get("tile") or "None"),
        ]
        for idx, (label, value) in enumerate(filters_map, start=3):
            ws_stats[f"A{idx}"] = label
            ws_stats[f"B{idx}"] = value

        # Report generation timestamp, boxed with filters
        ts_row = 3 + len(filters_map)
        ws_stats[f"A{ts_row}"] = "Generated at"
        ws_stats[f"B{ts_row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Apply light card-style fill and border to filters + generated-at block
        card_fill = PatternFill("solid", fgColor="F1F5F9")
        card_border_side = Side(style="thin", color="CBD5E1")
        card_border = Border(top=card_border_side, left=card_border_side, right=card_border_side, bottom=card_border_side)
        for row in range(2, ts_row + 1):
            for col in (1, 2):
                cell = ws_stats.cell(row=row, column=col)
                cell.fill = card_fill
                cell.border = card_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        # Bold filter labels
        for row in range(3, ts_row):
            ws_stats[f"A{row}"].font = Font(bold=True, size=10, color="111827")

        # Basic stats table
        stats_header_row = ts_row + 2
        ws_stats[f"A{stats_header_row}"] = "Metric"
        ws_stats[f"B{stats_header_row}"] = "Value"
        ws_stats[f"A{stats_header_row}"].font = ws_stats[f"B{stats_header_row}"].font = Font(bold=True)
        metrics = [
            ("Starlink Activated", stats.get("star_activated", 0)),
            ("Starlink Not Activated", stats.get("star_not_activated", 0)),
            ("Approval Accepted", stats.get("approval_accepted", 0)),
            ("Approval Pending/Blank", stats.get("approval_pending", 0)),
            ("Approval Decline/Other", stats.get("approval_decline", 0)),
            ("Calendar Sent", stats.get("calendar_sent", 0)),
            ("Calendar Invite Not Sent", stats.get("calendar_not_sent", 0)),
            ("S1 - Installed (Success)", stats.get("s1_success", 0)),
        ]
        stats_start = stats_header_row + 1
        for idx, (label, value) in enumerate(metrics, start=stats_start):
            ws_stats[f"A{idx}"] = label
            ws_stats[f"B{idx}"] = int(value)

        # Style stats block as a card
        stats_end = stats_start + len(metrics) - 1
        stats_card_fill = PatternFill("solid", fgColor="EEF2FF")
        for row in range(stats_header_row, stats_end + 1):
            for col in (1, 2):
                cell = ws_stats.cell(row=row, column=col)
                cell.border = card_border
                cell.fill = stats_card_fill if row > stats_header_row else header_fill
                if col == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
        # Grouping blank rows between logical sections
        # (already implied by metric ordering; no extra blank rows needed)

        # Starlink pie chart
        ws_stats["D2"] = "Starlink Status"
        ws_stats["D3"] = "Activated"
        ws_stats["E3"] = int(stats.get("star_activated", 0))
        ws_stats["D4"] = "Not Activated"
        ws_stats["E4"] = int(stats.get("star_not_activated", 0))

        star_pie = PieChart()
        star_pie.title = "Starlink Status"
        data = Reference(ws_stats, min_col=5, min_row=3, max_row=4)
        labels = Reference(ws_stats, min_col=4, min_row=3, max_row=4)
        star_pie.add_data(data, titles_from_data=False)
        star_pie.set_categories(labels)
        # Match dashboard colors: green (#22C55E) vs red (#EF4444)
        star_series = star_pie.series[0]
        dp_activated = DataPoint(idx=0)
        dp_activated.graphicalProperties.solidFill = "22C55E"
        dp_not = DataPoint(idx=1)
        dp_not.graphicalProperties.solidFill = "EF4444"
        star_series.dpt = [dp_activated, dp_not]
        # Show values and percentages on slices
        star_pie.dataLabels = DataLabelList()
        star_pie.dataLabels.showVal = True
        star_pie.dataLabels.showPercent = True
        star_pie.width = 10
        star_pie.height = 6
        ws_stats.add_chart(star_pie, "H2")

        # Approval pie chart
        ws_stats["D8"] = "Approval Status"
        ws_stats["D9"] = "Accepted"
        ws_stats["E9"] = int(stats.get("approval_accepted", 0))
        ws_stats["D10"] = "Pending/Blank"
        ws_stats["E10"] = int(stats.get("approval_pending", 0))
        ws_stats["D11"] = "Decline/Other"
        ws_stats["E11"] = int(stats.get("approval_decline", 0))

        appr_pie = PieChart()
        appr_pie.title = "Approval Status"
        data2 = Reference(ws_stats, min_col=5, min_row=9, max_row=11)
        labels2 = Reference(ws_stats, min_col=4, min_row=9, max_row=11)
        appr_pie.add_data(data2, titles_from_data=False)
        appr_pie.set_categories(labels2)
        # Match dashboard colors: green (#22C55E), yellow (#FACC15), red (#EF4444)
        appr_series = appr_pie.series[0]
        dp_acc = DataPoint(idx=0)
        dp_acc.graphicalProperties.solidFill = "22C55E"
        dp_pending = DataPoint(idx=1)
        dp_pending.graphicalProperties.solidFill = "FACC15"
        dp_decline = DataPoint(idx=2)
        dp_decline.graphicalProperties.solidFill = "EF4444"
        appr_series.dpt = [dp_acc, dp_pending, dp_decline]
        # Show values and percentages on slices
        appr_pie.dataLabels = DataLabelList()
        appr_pie.dataLabels.showVal = True
        appr_pie.dataLabels.showPercent = True
        appr_pie.width = 10
        appr_pie.height = 6
        ws_stats.add_chart(appr_pie, "H18")

    return wb


@app.route("/", defaults={"path": ""}, methods=["GET"])
@app.route("/<path:path>", methods=["GET"])
def index(path: str):
    # Main dashboard handler (also handles report download when ?download=xlsx)
    selected_region = request.args.get("region", "").strip() or None
    selected_schedule = request.args.get("schedule", "").strip() or None
    selected_installation = request.args.get("installation", "").strip() or None
    selected_tile = request.args.get("tile", "").strip() or None
    selected_lot = request.args.get("lot", "").strip() or None
    selected_search = request.args.get("search", "").strip() or None
    include_unscheduled = request.args.get("full", "") == "1"

    (
        rows,
        region_options,
        schedule_options,
        installation_options,
        stats,
      ) = get_table_data(
          selected_region,
          selected_schedule,
          selected_installation,
          selected_tile,
          selected_lot,
          include_unscheduled=include_unscheduled,
          selected_search=selected_search,
    )

    # If download flag is present, stream XLSX instead of HTML
    if request.args.get("download") == "xlsx":
        selected_columns = request.args.getlist("col")
        include_stats = request.args.get("include_stats", "1") == "1"
        filters = {
            "region": selected_region,
            "schedule": selected_schedule,
            "installation": selected_installation,
            "tile": selected_tile,
            "lot": selected_lot,
        }
        wb = _build_workbook(rows, stats, selected_columns, include_stats, filters)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        stamp = datetime.now().strftime("%Y%m%d-%H%M")
        lot_tag = ""
        if selected_lot:
            lot_tag = "-" + selected_lot.lower().replace(" ", "").replace("#", "")
        filename = f"monitoring-report{lot_tag}-{stamp}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    show_report = request.args.get("report", "") == "1"

    return render_template_string(
        TEMPLATE,
        rows=rows,
        region_options=region_options,
        schedule_options=schedule_options,
          installation_options=installation_options,
          selected_region=selected_region or "",
          selected_schedule=selected_schedule or "",
          selected_installation=selected_installation or "",
          selected_tile=selected_tile or "",
          selected_lot=selected_lot or "",
          selected_search=selected_search or "",
          stats=stats,
        show_report=show_report,
        include_unscheduled=include_unscheduled,
        last_updated=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )


# On Vercel, the `app` object is used as the WSGI entrypoint.
